import openpyxl
import requests
import time
import random
wb = openpyxl.load_workbook('savedrecs.xlsx')
wb_savedrecs = wb['savedrecs']
max_col = wb_savedrecs.max_column
max_row = wb_savedrecs.max_row
doi = []
for col in range(1,max_col+1):
	if wb_savedrecs.cell(row=1,column=col).value == 'DOI':
		doicol = col
		break
for rowi in range(2,max_row+1):
	if wb_savedrecs.cell(row=rowi,column=doicol).value != '':
		doi.append(wb_savedrecs.cell(row=rowi,column=doicol).value)
# with open('doi.txt','a') as fw:
# 	for item in doi:
# 		fw.write(item+'\n')
# doi = open("doi.txt",encoding = "utf-8").readlines()
#doi = ['10.1103/PhysRevB.103.024516','10.1088/1361-6668/aafa88','10.1109/TASC.2021.3066130']
url = 'https://sci-hub.se/'
headers = {
	'referer': 'https://sci-hub.se/',
	'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'
}
def downpaper(url,doi,headers):
	text = requests.get(url+doi,headers).text
	#print(text)
	downurl = ''
	savei = text.find('下载')
	if savei != -1:
		while True:
			savei = savei - 1
			if text[savei] == 'f' and text[savei-1] == 'e' and text[savei-2] == 'r' and text[savei-3] == 'h':
				hreflo = savei + 3
				hrefi = hreflo
				while True:
					hrefi += 1
					if text[hrefi] == '>':
						hrefhi = hrefi - 2
						break
				for i in range(hreflo,hrefhi):
					downurl += text[i]
				break
	else:
		with open('notfind.txt','a') as fw:
			fw.write(doi+'\n')
		return
	year = ''
	title = ''
	if text.find('(20') != -1:
		yeari = text.find('(20')
		flag = 0
		while True:
			yeari += 1
			if '0'<=text[yeari]<='9' and flag < 4:
				flag += 1
				year += text[yeari]
			if flag == 4:
				titlei = yeari
				break
		flag = 0
		while True:
			titlei += 1
			if flag == 0 and text[titlei] == '.':
				flag += 1
				while True:
					titlei += 1
					if text[titlei] != ' ':
						titlelo = titlei
						if text[titlei]=='<' and text[titlei+2]=='>':
							titlelo = titlei + 3
						break
			if text[titlei] == '.':
				titlehi = titlei
				for i in range(titlelo,titlehi):
					if text[i]!='\n':
						title += text[i]
				title = year +'-'+ title
				break
	else:
		title = ''
		newtitlei = text.find('.pdf')
		newtitlehi = newtitlei
		while True:
			newtitlei = newtitlei - 1
			if text[newtitlei] == '/':
				newtitlelo = newtitlei + 1
				for i in range(newtitlelo,newtitlehi):
					title += text[i]
				break
	print(doi)
	with open('title.txt','a') as fw:
		fw.write(doi+'\n')
	print(title)
	with open('title.txt','a') as fw:
		fw.write(title+'\n\n')
	tsp = title.split(':')
	tnew = ''
	for i in range(len(tsp)):
		if i == 0:
			tnew += tsp[i]
		else:
			tnew = tnew +'..'+tsp[i]
	title = tnew
	if downurl[0] == '/' and downurl[1] == '/':
		downurl = 'http:' + downurl
	else:
		downurl = 'http://sci-hub.se' + downurl
	print(downurl+'\n')
	try:
		r = requests.get(downurl,headers)
		with open(title+'.pdf','wb') as f:
			f.write(r.content)
	except:
		print('下载出错')
		with open('error.txt','a') as fw:
			fw.write(doi+'\n')
for doii in doi:
	downpaper(url,doii.split()[0],headers)
	time.sleep(2+random.random())