from bs4 import BeautifulSoup
import requests
import openpyxl
import time
import random
import re

wb = openpyxl.load_workbook('savedrecs.xlsx')
wb_savedrecs = wb['savedrecs']
max_col = wb_savedrecs.max_column
max_row = wb_savedrecs.max_row
doi = []
for col in range(1,max_col+1):
    if wb_savedrecs.cell(row=1,column=col).value == 'DOI':
        doicol = col
        break
for rowi in range(2,max_row+1):
    if wb_savedrecs.cell(row=rowi,column=doicol).value != '':
        doi.append(wb_savedrecs.cell(row=rowi,column=doicol).value)

url = 'https://www.tesble.com/'
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
    'Priority': 'u=0, i',
    'Sec-Ch-Ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Microsoft Edge";v="126"',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0'
}

dl_headers = {
    # ':authority': 'sci.bban.top',
    # ':method': 'GET',
    # ':path': '/pdf/10.1088/1361-6668%252F30%252F3%252F033004.pdf',
    # ':scheme': 'https',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
    'Priority': 'u=0, i',
    'Referer': 'https://www.tesble.com/',
    'Sec-Ch-Ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Microsoft Edge";v="126"',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Sec-Fetch-Dest': 'embed',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'cross-site',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0'
}


def downpaper(url,doi,headers):
    print(f'Downloading paper with doi: {doi}')
    response = requests.post(url, headers=headers, data={'request': doi})
    text = response.text
    # print(text)
    downurl = ''

    soup = BeautifulSoup(text, 'html.parser')
    embed_tag = soup.find('embed', {'id': 'pdf'})
    if embed_tag and 'src' in embed_tag.attrs:
        downurl = embed_tag['src']
        print(downurl)
        filename = re.sub(r'[\\/*?:"<>|]', "_", doi)
        try:
            r = requests.get(downurl, headers=dl_headers)
            r.raise_for_status()
            with open(f'{filename}.pdf', 'wb') as f:
                f.write(r.content)
            print(f'PDF downloaded and saved as {doi}.pdf')
        except requests.exceptions.RequestException as e:
            print(f'Error downloading PDF: {e}')
            with open('error.txt', 'a') as fw:
                fw.write(f'{filename} - PDF download error: {e}\n')

for doii in doi:
    downpaper(url,doii.split()[0],headers)
    time.sleep(2+random.random())